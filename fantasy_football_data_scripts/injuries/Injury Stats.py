import nfl_data_py as nfl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import re

# Define the manual mapping for player names (case insensitive)
manual_mapping = {
    "hollywood brown": "Marquise Brown",
    "will fuller v": "Will Fuller",
    "jeff wilson": "Jeffery Wilson",
    "willie snead iv": "Willie Snead",
    "charles johnson": "Charles D Johnson",
    "kenneth barber": "Peyton Barber",
    "rodney smith": "Rod Smith",
    "bisi johnson": "Olabisi Johnson",
    "chris herndon": "Christopher Herndon",
    "scotty miller": "Scott Miller",
    "trenton richardson": "Trent Richardson"
}

# Function to clean player names
def clean_name(name):
    # Replace specific character sequences
    name = re.sub(r'[èéêëÈÉÊË]', 'e', name)

    pattern = r"[.\-']|(\bjr\b\.?)|(\bsr\b\.?)|(\bII\b)|(\bIII\b)"
    cleaned_name = re.sub(pattern, '', name, flags=re.IGNORECASE).strip()
    capitalized_name = ' '.join([word.capitalize() for word in cleaned_name.split()])
    return capitalized_name

# Function to apply manual mapping
def apply_manual_mapping(name):
    return manual_mapping.get(name.lower(), name)

def calculate_cumulative_week(row):
    if row['season'] == 2013:
        return row['week'] + (row['season'] - 2014) * 14
    elif row['season'] <= 2020:
        return row['week'] + (row['season'] - 2014) * 16 + 14
    else:
        return row['week'] + (row['season'] - 2014) * 17 + 14

def add_custom_columns(df):
    if 'season' not in df.columns:
        raise KeyError("The 'season' column is missing from the DataFrame.")
    df['Cumulative Week'] = df.apply(calculate_cumulative_week, axis=1)
    df['playerweek'] = df.apply(lambda row: f"{row['full_name']}{int(row['Cumulative Week'])}".replace(" ", ""), axis=1)
    df['playeryear'] = df.apply(lambda row: f"{row['full_name']}{int(row['season'])}".replace(" ", ""), axis=1)
    return df

def download_injury_data(years, weeks, output_file):
    # Load injury data for the specified years
    injury_data = nfl.import_injuries(years=years)

    if weeks == [0]:
        if all(year <= 2020 for year in years):
            weeks = list(range(1, 17))
        else:
            weeks = list(range(1, 18))

    injury_data = injury_data[injury_data['week'].isin(weeks)]

    # Convert timezone-aware datetimes to timezone-unaware
    for col in injury_data.select_dtypes(include=['datetimetz']).columns:
        injury_data[col] = injury_data[col].dt.tz_localize(None)

    # Rename teams
    team_replacements = {
        'STL': 'LAR',
        'SD': 'LAC',
        'LA': 'LAR',
        'OAK': 'LV'
    }
    injury_data['team'] = injury_data['team'].replace(team_replacements)

    # Clean and map player names
    injury_data['full_name'] = injury_data['full_name'].apply(clean_name).apply(apply_manual_mapping)

    # Add custom columns
    injury_data = add_custom_columns(injury_data)

    # Save the injury data to a CSV file
    injury_data.to_csv(output_file, index=False)
    print(f"Injury data for {years} and weeks {weeks} saved to '{output_file}'!")

    # Save the injury data to an Excel file and format as a table
    excel_file = output_file.replace('.csv', '.xlsx')
    injury_data.to_excel(excel_file, index=False, sheet_name='Injury Data')

    wb = load_workbook(excel_file)
    ws = wb['Injury Data']

    table_name = "InjuryDataTable"
    min_col, min_row = 1, 1
    max_col = get_column_letter(len(injury_data.columns))
    max_row = ws.max_row
    tab = Table(displayName=table_name, ref=f"A{min_row}:{max_col}{max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(excel_file)
    print(f"Excel file with formatted table has been created at {excel_file}")

    # Remove the CSV file
    os.remove(output_file)
    print(f"CSV file '{output_file}' has been removed.")

if __name__ == "__main__":
    season_year = int(input("Enter the NFL season year (0 for all seasons from 2014 to present): "))
    week_number = int(input("Enter the week number (0 for all weeks): "))

    if season_year == 0:
        years = list(range(2014, pd.Timestamp.now().year + 1))
    else:
        years = [season_year]

    if week_number == 0:
        weeks = [0]
    else:
        weeks = [week_number]

    base_dir = r'C:\Users\joeye\OneDrive\Desktop\kmffl\Adin\Scripts\Sheet 2.0'
    output_file = os.path.join(base_dir, f"injury_data_{season_year}_{week_number}.csv")
    download_injury_data(years, weeks, output_file)