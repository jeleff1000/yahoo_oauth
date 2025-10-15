import os
import re
import requests
import time
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from xml.etree import ElementTree as ET
from yahoo_oauth import OAuth2
import yahoo_fantasy_api as yfa
import pandas as pd

class DraftResult:
    def __init__(self, extracted_data):
        self.cost = extracted_data.get("cost", "N/A")
        self.pick = extracted_data.get("pick", "N/A")
        self.round = extracted_data.get("round", "N/A")
        self.team_key = extracted_data.get("team_key", "N/A")
        self.player_id = extracted_data.get("player_id", "")
        self.year = None

def fetch_draft_data(oauth, league_id):
    league = yfa.League(oauth, league_id)
    draft_results = league.draft_results()
    return [DraftResult(result) for result in draft_results]

def fetch_team_and_player_mappings(oauth, league_id):
    team_key_to_manager = {}
    player_id_to_name = {}
    statMapping = {}

    # Fetch stat mappings
    url = f"https://fantasysports.yahooapis.com/fantasy/v2/league/{league_id}/settings"
    max_retries = 40
    backoff_factor = 3

    for attempt in range(max_retries):
        try:
            r = oauth.session.get(url)
            r.raise_for_status()  # Raise an HTTPError for bad responses
            xmlstring = re.sub(' xmlns="[^"]+"', '', r.text, count=1)
            if not xmlstring.strip():  # Check if the XML string is empty
                raise ValueError("Empty XML response")
            root = ET.fromstring(xmlstring)
            break  # Exit the retry loop if successful
        except (requests.ConnectionError, requests.Timeout, requests.exceptions.HTTPError) as e:
            if attempt < max_retries - 1:
                sleep_time = backoff_factor * (2 ** attempt)
                print(f"Connection error: {e}. Retrying in {sleep_time} seconds...")
                time.sleep(sleep_time)
            else:
                print("Max retries reached. Exiting.")
                raise
        except ET.ParseError as e:
            print(f"XML parse error: {e}. Exiting.")
            raise
        except ValueError as e:
            print(f"Error: {e}. Exiting.")
            raise

    for stat in root.findall("league/settings/stat_categories/stats/stat"):
        stat_id = stat.find("stat_id").text
        display_name = stat.find("display_name").text
        statMapping[stat_id] = display_name

    # Fetch team and player mappings for Week 1
    for i in range(1, 11):
        url = f"https://fantasysports.yahooapis.com/fantasy/v2/team/{league_id}.t.{i}/roster;week=1/players/stats"
        r = oauth.session.get(url)
        xmlstring = r.text
        if not xmlstring.strip():
            continue
        xmlstring = re.sub(' xmlns="[^"]+"', '', xmlstring, count=1)
        root = ET.fromstring(xmlstring)

        manager_element = root.find("team/managers/manager/nickname")
        if manager_element is not None:
            nickname = manager_element.text
        else:
            nickname = "Unknown"

        team_key = root.find("team/team_key").text
        team_key_to_manager[team_key] = nickname  # Use nickname directly

        for player in root.findall("team/roster/players/player"):
            player_id = player.find("player_id").text
            if player_id not in player_id_to_name:
                name = player.find("name/full").text
                player_id_to_name[player_id] = name

    # Fetch all players with pagination
    start = 0
    while True:
        url = f"https://fantasysports.yahooapis.com/fantasy/v2/league/{league_id}/players;start={start}"
        r = oauth.session.get(url)
        xmlstring = r.text
        if not xmlstring.strip():
            break
        xmlstring = re.sub(' xmlns="[^"]+"', '', xmlstring, count=1)
        root = ET.fromstring(xmlstring)

        players = root.findall("league/players/player")
        if not players:
            break

        for player in players:
            player_id = player.find("player_id").text
            if player_id not in player_id_to_name:
                name = player.find("name/full").text
                player_id_to_name[player_id] = name

        start += len(players)

    return team_key_to_manager, player_id_to_name, statMapping

def fetch_draft_analysis(oauth, league_id, year):
    """Fetches additional draft analysis information."""
    draft_analysis = []

    start = 0
    while True:
        url = f"https://fantasysports.yahooapis.com/fantasy/v2/league/{league_id}/players;start={start}/draft_analysis"
        r = oauth.session.get(url)
        xmlstring = r.text
        if not xmlstring.strip():
            break

        xmlstring = re.sub(' xmlns="[^"]+"', '', xmlstring, count=1)
        try:
            root = ET.fromstring(xmlstring)
        except ET.ParseError as e:
            print(f"XML parse error: {e}. Skipping this batch.")
            start += 25  # Assuming each batch contains 25 players
            continue

        players = root.findall("league/players/player")
        if not players:
            break

        for player in players:
            player_data = {
                "year": year,
                "player_key": player.find("player_key").text,
                "player_id": player.find("player_id").text,
                "name_full": player.find("name/full").text,
                "primary_position": player.find("primary_position").text,
                "average_pick": player.find("draft_analysis/average_pick").text,
                "average_round": player.find("draft_analysis/average_round").text,
                "average_cost": player.find("draft_analysis/average_cost").text,
                "percent_drafted": player.find("draft_analysis/percent_drafted").text,
                "preseason_average_pick": player.find("draft_analysis/preseason_average_pick").text,
                "preseason_average_round": player.find("draft_analysis/preseason_average_round").text,
                "preseason_average_cost": player.find("draft_analysis/preseason_average_cost").text,
                "preseason_percent_drafted": player.find("draft_analysis/preseason_percent_drafted").text,
                "is_keeper_status": player.find("is_keeper/status").text if player.find("is_keeper/status") is not None else "",
                "is_keeper_cost": player.find("is_keeper/cost").text if player.find("is_keeper/cost") is not None else ""
            }
            draft_analysis.append(player_data)

        start += len(players)

    return draft_analysis

# Manager name mapping
managerNameToStatName = {
    "Jason": "Kardon",
    "Marc": "Leeb",
    "Gavi": "Gavi",
    "Adin": "Adin",
    "Tani": "Tani",
    "Yaacov": "Yaacov",
    "joey": "Eleff",
    "Ezra": "Newman",
    "Daniel": "Daniel",
    "Jesse": "Jesse"
}

def write_draft_data_to_csv(draft_data, file_path, team_key_to_manager):
    with open(file_path, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Year", "Pick", "Round", "Team Key", "Team Manager", "Player ID", "Cost"])
        for result in draft_data:
            team_manager = team_key_to_manager.get(result.team_key, "N/A")
            # Map the manager name
            mapped_manager = managerNameToStatName.get(team_manager, team_manager)
            writer.writerow([result.year, result.pick, result.round, result.team_key, mapped_manager, result.player_id, result.cost])

def write_draft_analysis_to_csv(draft_analysis, file_path):
    with open(file_path, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([
            "Year", "Player Key", "Player ID", "Name Full", "Primary Position", "Average Pick", "Average Round", "Average Cost",
            "Percent Drafted", "Preseason Average Pick", "Preseason Average Round", "Preseason Average Cost", "Preseason Percent Drafted",
            "Is Keeper Status", "Is Keeper Cost"
        ])
        for analysis in draft_analysis:
            writer.writerow([
                analysis["year"], analysis["player_key"], analysis["player_id"], analysis["name_full"], analysis["primary_position"],
                analysis["average_pick"], analysis["average_round"], analysis["average_cost"], analysis["percent_drafted"],
                analysis["preseason_average_pick"], analysis["preseason_average_round"], analysis["preseason_average_cost"],
                analysis["preseason_percent_drafted"], analysis["is_keeper_status"], analysis["is_keeper_cost"]
            ])

def merge_csvs(draft_data_file, draft_analysis_file, merged_file, year):
    draft_data_dict = {}
    with open(draft_data_file, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            draft_data_dict[row["Player ID"]] = row

    with open(draft_analysis_file, 'r') as file:
        reader = csv.DictReader(file)
        draft_analysis_data = list(reader)

    with open(merged_file, 'a', newline='') as file:
        writer = csv.writer(file)
        for analysis in draft_analysis_data:
            player_id = analysis["Player ID"]
            draft_data = draft_data_dict.get(player_id, {})
            cost_str = draft_data.get("Cost", "0").replace("N/A", "0")
            average_cost_str = analysis.get("Average Cost", "0").replace("N/A", "0")

            try:
                cost = float(cost_str)
                average_cost = float(average_cost_str)
                if cost == 0 or average_cost == 0:
                    savings = "N/A"
                else:
                    savings = average_cost - cost
            except ValueError:
                savings = "N/A"

            playeryear = f"{analysis.get('Name Full').strip()}{year}".replace(' ', '')
            manager_name = draft_data.get("Team Manager", "N/A")
            mapped_manager = managerNameToStatName.get(manager_name, manager_name)
            manageryear = f"{mapped_manager.strip()}{year}".replace(' ', '')

            writer.writerow([
                year, draft_data.get("Pick", "N/A"), draft_data.get("Round", "N/A"), draft_data.get("Team Key", "N/A"),
                manager_name, player_id, cost_str,
                analysis.get("Name Full", ""), analysis.get("Primary Position", ""), analysis.get("Average Pick", ""),
                analysis.get("Average Round", ""), average_cost_str, analysis.get("Percent Drafted", ""),
                analysis.get("Is Keeper Status", ""), analysis.get("Is Keeper Cost", ""), savings, playeryear, manageryear
            ])

def create_excel_table(merged_file, excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft History"

    headers = [
        "Year", "Pick", "Round", "Team_Key", "Manager", "Player Key", "Player ID", "Name Full", "Primary Position", "Average Pick", "Average Round", "Average Cost",
        "Percent Drafted", "Preseason Average Pick", "Preseason Average Round", "Preseason Average Cost", "Preseason Percent Drafted",
        "Is Keeper Status", "Is Keeper Cost"
    ]
    ws.append(headers)

    with open(merged_file, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            ws.append(row)

    tab = Table(displayName="MergedDataTable", ref=f"A1:{chr(65 + len(headers) - 1)}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Convert numbers stored as text to numbers
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.isdigit():
                cell.value = int(cell.value)
            elif isinstance(cell.value, str):
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    pass

    wb.save(excel_file)

def fetch_all_years_data(oauth, gm, output_dir):
    current_year = datetime.now().year
    merged_file = os.path.join(output_dir, 'merged_draft_data_0.csv')

    # Create the merged file and write the header
    with open(merged_file, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([
            "Year", "Pick", "Round", "Team Key", "Team Manager", "Player ID", "Cost",
            "Name Full", "Primary Position", "Average Pick", "Average Round", "Average Cost",
            "Percent Drafted", "Is Keeper Status", "Is Keeper Cost", "Savings", "playeryear", "manageryear"
        ])

    for year in range(2014, current_year + 1):
        if not oauth.token_is_valid():
            oauth.refresh_access_token()

        try:
            league_ids = gm.league_ids(year=year)
        except RuntimeError as e:
            print(f"Error fetching league IDs for year {year}: {e}")
            continue

        if not league_ids:
            print(f"No league IDs found for year {year}. Skipping.")
            continue
        league_id = league_ids[-1]

        team_key_to_manager, player_id_to_name, statMapping = fetch_team_and_player_mappings(oauth, league_id)
        draft_data = fetch_draft_data(oauth, league_id)
        for data in draft_data:
            data.year = year

        draft_analysis = fetch_draft_analysis(oauth, league_id, year)

        draft_data_file = os.path.join(output_dir, f'draft_data_{year}.csv')
        draft_analysis_file = os.path.join(output_dir, f'draft_analysis_{year}.csv')

        write_draft_data_to_csv(draft_data, draft_data_file, team_key_to_manager)
        write_draft_analysis_to_csv(draft_analysis, draft_analysis_file)
        merge_csvs(draft_data_file, draft_analysis_file, merged_file, year)

        # Delete the intermediate files
        os.remove(draft_data_file)
        os.remove(draft_analysis_file)

    return merged_file

def main():
    oauth = OAuth2(None, None, from_file='C:\\Users\\joeye\\OneDrive\\Desktop\\kmffl\\Adin\\Oauth.json')
    if not oauth.token_is_valid():
        oauth.refresh_access_token()

    gm = yfa.Game(oauth, 'nfl')
    year = int(input("Select the year to get data for (type 0 for all years since 2014): "))
    week_input = input("Select the week to get data for (type 0 for the whole year): ")

    output_dir = 'C:\\Users\\joeye\\OneDrive\\Desktop\\kmffl\\Adin\\Scripts\\Sheet 2.0'
    os.makedirs(output_dir, exist_ok=True)

    if year == 0:
        combined_file = fetch_all_years_data(oauth, gm, output_dir)
        excel_file = os.path.join(output_dir, 'combined_draft_data.xlsx')
        create_excel_table(combined_file, excel_file)
        os.remove(combined_file)  # Delete the combined CSV file
        print(f"Combined draft data for all years has been written to {combined_file}")
        print(f"Excel file with formatted table has been created at {excel_file}")
    else:
        league_ids = gm.league_ids(year=year)
        if not league_ids:
            print(f"No league IDs found for year {year}. Exiting.")
            return

        league_id = league_ids[-1]

        team_key_to_manager, player_id_to_name, statMapping = fetch_team_and_player_mappings(oauth, league_id)
        draft_data = fetch_draft_data(oauth, league_id)
        for data in draft_data:
            data.year = year

        draft_analysis = fetch_draft_analysis(oauth, league_id, year)

        draft_data_file = os.path.join(output_dir, f'draft_data_{year}.csv')
        draft_analysis_file = os.path.join(output_dir, f'draft_analysis_{year}.csv')
        merged_file = os.path.join(output_dir, f'merged_draft_data_{year}.csv')
        excel_file = os.path.join(output_dir, f'merged_draft_data_{year}.xlsx')

        write_draft_data_to_csv(draft_data, draft_data_file, team_key_to_manager)
        write_draft_analysis_to_csv(draft_analysis, draft_analysis_file)
        merge_csvs(draft_data_file, draft_analysis_file, merged_file, year)
        create_excel_table(merged_file, excel_file)

        os.remove(draft_data_file)  # Delete the draft data CSV file
        os.remove(draft_analysis_file)  # Delete the draft analysis CSV file
        os.remove(merged_file)  # Delete the merged CSV file

        print(f"Draft data for {year} has been written to {draft_data_file}")
        print(f"Draft analysis data for {year} has been written to {draft_analysis_file}")
        print(f"Merged draft data for {year} has been written to {merged_file}")
        print(f"Excel file with formatted table has been created at {excel_file}")

    excel_file_name = os.path.join(output_dir, 'draft_history.xlsx')
    print(excel_file_name)

if __name__ == "__main__":
    main()