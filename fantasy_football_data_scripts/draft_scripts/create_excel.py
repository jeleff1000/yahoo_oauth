import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel_table(merged_file, excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft History"

    headers = [
        "Year", "Pick", "Round", "Team Key", "Team Manager", "Player ID", "Cost", "Name Full", "Primary Position",
        "Average Pick", "Average Round", "Average Cost", "Percent Drafted", "Is Keeper Status", "Is Keeper Cost",
        "Savings", "PlayerYear", "ManagerYear", "Editorial Team Abbr"
    ]
    ws.append(headers)

    def convert_to_numeric(value):
        try:
            return int(value)
        except ValueError:
            try:
                return float(value)
            except ValueError:
                return value

    with open(merged_file, 'r') as file:
        reader = csv.reader(file)
        next(reader)  # Skip the original header row
        for row in reader:
            row = [convert_to_numeric(cell) for cell in row]
            ws.append(row)

    # Check for duplicates in PlayerYear and update values
    player_years = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers), values_only=True), start=2):
        player_year = row[16]  # PlayerYear column index
        name_full = row[7].replace(" ", "")  # Name Full column index
        primary_position = row[8].replace(" ", "")  # Primary Position column index
        year = row[0]  # Year column index
        editorial_team_abbr = row[18].replace(" ", "")  # Editorial Team Abbr column index

        if player_year in player_years:
            new_player_year = f"{name_full}{primary_position}{year}"
            if new_player_year in player_years:
                new_player_year = f"{name_full}{primary_position}{editorial_team_abbr}{year}"
            player_years[new_player_year] = player_years.get(new_player_year, 0) + 1
            ws.cell(row=row_idx, column=17, value=new_player_year)  # Update PlayerYear cell
        else:
            player_years[player_year] = 1

    tab = Table(displayName="MergedDataTable", ref=f"A1:{chr(65 + len(headers) - 1)}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(excel_file)